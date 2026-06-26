from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from openpyxl import load_workbook
from io import BytesIO
import json
import os
from werkzeug.utils import secure_filename
import requests
import math
from dateutil import parser as dateutil_parser

app = Flask(__name__)

# CORS configuration
# In production, set CORS_ORIGINS env var to your frontend domain(s), comma-separated.
# Example: CORS_ORIGINS=https://importer.yourcompany.com
_cors_origins = os.environ.get('CORS_ORIGINS', 'http://localhost:5173,http://127.0.0.1:5173')
ALLOWED_ORIGINS = [o.strip() for o in _cors_origins.split(',') if o.strip()]
CORS(app, resources={
    r"/*": {
        "origins": ALLOWED_ORIGINS,
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "OrcanosAPIKey", "X-Orcanos-Domain"],
        "supports_credentials": True
    }
})

ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = 'uploads'

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def serialize_value(obj):
    """Safe JSON serialization for NaN, None, infinite values"""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
    return str(obj)

def _response_is_json(response):
    content_type = response.headers.get('Content-Type', '')
    return 'application/json' in content_type.lower()

def parse_orcanos_json(response):
    """
    Parse Orcanos API response body only when Content-Type is application/json.
    Returns (data, error_message). error_message is None on success.
    """
    text = response.text or ''
    if not text.strip():
        return None, "Server returned empty response. Please check your domain URL."
    if not _response_is_json(response):
        return None, (
            "Server returned unexpected response. Please check your domain URL is correct. "
            f"Received: {text[:200]}"
        )
    try:
        return response.json(), None
    except (json.JSONDecodeError, requests.exceptions.JSONDecodeError) as e:
        return None, f"Invalid JSON response: {str(e)}"

@app.route('/api/verify-auth', methods=['POST'])
@app.route('/verify-auth', methods=['POST'])
def verify_auth():
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({"valid": False, "error": "No data provided"}), 400

        domain = payload.get('domain', '')
        headers = payload.get('headers', {})

        if not domain:
            return jsonify({"valid": False, "error": "Domain is required"}), 400

        url = f"https://{domain}/api/v2/Json/QW_Login"
        response = requests.post(url, headers=headers, timeout=30)

        # Handle 401/403 status codes
        if response.status_code in (401, 403):
            return jsonify({"valid": False, "error": "Access denied. Please check your credentials."})

        # If response is not JSON (HTML page, bad domain, etc.)
        response_data, parse_error = parse_orcanos_json(response)
        if parse_error:
            return jsonify({"valid": False, "error": "Cannot reach Orcanos. Please check your domain is correct."})

        if response_data.get('IsSuccess'):
            projects = response_data.get('Data', {}).get('Projects', {}).get('Project', [])
            return jsonify({"valid": True, "projectsList": projects})

        # Check for SSO / auth failure messages
        message = response_data.get('Message', '')
        message_lower = message.lower() if message else ''
        if 'sso' in message_lower or 'try using sso' in message_lower \
                or 'authentication' in message_lower or 'login' in message_lower \
                or 'password' in message_lower or 'credential' in message_lower \
                or not message:
            return jsonify({"valid": False, "error": "Incorrect username or password. Please try again."})

        # Any other Orcanos-returned message — still keep it clean
        return jsonify({"valid": False, "error": "Incorrect username or password. Please try again."})

    except requests.exceptions.Timeout:
        return jsonify({"valid": False, "error": "Connection timed out. Please check your domain and try again."})
    except requests.exceptions.ConnectionError:
        return jsonify({"valid": False, "error": "Cannot reach Orcanos. Please check your domain is correct."})
    except requests.exceptions.RequestException:
        return jsonify({"valid": False, "error": "Cannot reach Orcanos. Please check your domain is correct."})
    except Exception:
        return jsonify({"valid": False, "error": "Something went wrong. Please try again."}), 500

@app.route('/ping', methods=['GET'])
def ping():
    try:
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": f"Error in ping: {str(e)}"}), 500

@app.route('/api/get-item-fields', methods=['POST'])
@app.route('/get-item-fields', methods=['POST'])
def get_item_fields():
    try:
        payload = request.get_json() or {}
        
        # Extract inputs
        item_type = payload.get('item_type', '')
        project_id = payload.get('project_id')
        major_version = payload.get('major_version')
        minor_version = payload.get('minor_version')
        
        # Extract credentials from payload or request headers
        domain = payload.get('domain')
        if not domain:
            domain = request.headers.get('X-Orcanos-Domain') or request.headers.get('Domain')
            
        # Headers extraction (OrcanosAPIKey, Authorization)
        headers = {}
        
        # Check request headers
        orcanos_api_key = request.headers.get('OrcanosAPIKey')
        authorization = request.headers.get('Authorization')
        
        if orcanos_api_key:
            headers['OrcanosAPIKey'] = orcanos_api_key
        if authorization:
            headers['Authorization'] = authorization
            
        # Check payload/body headers if sent from front-end state
        payload_headers = payload.get('headers') or {}
        if isinstance(payload_headers, dict):
            for k, v in payload_headers.items():
                if k.lower() == 'orcanosapikey' and not orcanos_api_key:
                    headers['OrcanosAPIKey'] = v
                elif k.lower() == 'authorization' and not authorization:
                    headers['Authorization'] = v
                    
        # Ensure Content-Type header is JSON
        headers['Content-Type'] = 'application/json'
            
        url = f"https://{domain}/api/v2/Json/QW_Get_Item_Add_Edit"
        orcanos_payload = {
            "Item_Type": item_type,
            "Project_id": project_id,
            "Major_Version": major_version,
            "Minor_Version": minor_version
        }
        
        try:
            response = requests.post(url, json=orcanos_payload, headers=headers, timeout=30)

            response_data, parse_error = parse_orcanos_json(response)

            # If parse succeeded and Orcanos reports success, return as-is
            if not parse_error:
                is_success = response_data.get('IsSuccess')
                http_code = response_data.get('HttpCode')
                has_fields = (
                    isinstance(response_data.get('Data'), dict) and
                    isinstance(response_data['Data'].get('field'), list) and
                    len(response_data['Data']['field']) > 0
                )

                if is_success and http_code == 200 and has_fields:
                    return jsonify(response_data)

            return jsonify({"error": "No fields returned. Please check your Project ID, Item Type, and Version."}), 400

        except requests.exceptions.Timeout:
            return jsonify({"error": "Connection timed out. Please check your domain."}), 400
        except requests.exceptions.ConnectionError:
            return jsonify({"error": "Could not connect to Orcanos. Please check your domain."}), 400
        except Exception as e:
            return jsonify({"error": f"Unexpected error: {str(e)}"}), 500
            
    except Exception as e:
        return jsonify({"error": f"Internal server error: {str(e)}", "Message": f"Internal server error: {str(e)}"}), 500

@app.route('/api/upload', methods=['POST'])
@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        if not allowed_file(file.filename):
            return jsonify({"error": "Please upload a valid Excel (.xlsx) file"}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        wb = load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        main_sheet_name = request.form.get('mainSheet', '').strip()
        steps_sheet_name = request.form.get('stepsSheet', '').strip()

        def read_sheet(ws):
            """Read a worksheet and return (headers, all_rows, preview_rows, row_count)."""
            hdrs = []
            for cell in ws[1]:
                hdrs.append(cell.value if cell.value is not None else '')
            all_rows = []
            preview = []
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_values = [cell.value for cell in row]
                if all(v is None for v in row_values):
                    continue
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    header = hdrs[col_idx] if col_idx < len(hdrs) else f"Column {col_idx + 1}"
                    row_dict[header] = cell.value
                all_rows.append(row_dict)
                if len(preview) < 5:
                    preview.append(row_dict)
                count += 1
            return hdrs, all_rows, preview, count

        # Determine which worksheet to use for main data
        if main_sheet_name and main_sheet_name in sheet_names:
            ws_main = wb[main_sheet_name]
        else:
            # Mode 1 (no sheet selected yet) or non-Test-Case flow: use active sheet
            ws_main = wb.active

        main_headers, main_rows, main_preview, main_count = read_sheet(ws_main)

        # Read steps sheet if specified
        steps_headers = None
        steps_data = None
        if steps_sheet_name and steps_sheet_name not in ('None', '') and steps_sheet_name in sheet_names:
            ws_steps = wb[steps_sheet_name]
            steps_headers, steps_data, _, _ = read_sheet(ws_steps)

        wb.close()
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception as cleanup_err:
            print(f"Warning: Could not delete temp file {filepath}: {str(cleanup_err)}")

        response_body = {
            "headers": main_headers,
            "preview": main_preview,
            "data": main_rows,
            "totalRows": main_count,
            "sheetNames": sheet_names
        }
        if steps_headers is not None:
            response_body["stepsHeaders"] = steps_headers
        if steps_data is not None:
            response_body["stepsData"] = steps_data

        return Response(
            json.dumps(response_body, default=serialize_value),
            mimetype='application/json',
            status=200
        )
    except Exception as e:
        return jsonify({"error": f"Error processing file: {str(e)}"}), 500

def validate_row(api_body, mandatory_fields=None):
    if mandatory_fields is None:
        mandatory_fields = []
        
    reasons = []
    
    # 1. Check mandatory fields are present and not empty
    for field in mandatory_fields:
        val = api_body.get(field)
        if val is None or (isinstance(val, str) and val.strip() == ''):
            reasons.append(f"{field} is required")

   

    # 3. Object_Name max length
    obj_name = api_body.get('Object_Name')
    if obj_name is not None and isinstance(obj_name, str) and len(obj_name) > 255:
        reasons.append("Object_Name must not exceed 255 characters")

    # 4. Description must not be empty or just whitespace
    desc = api_body.get('Description')
    if desc is not None and isinstance(desc, str) and desc.strip() == '':
        # Only add if not already caught by mandatory check
        if "Description is required" not in reasons:
            reasons.append("Description must not be empty or just whitespace")

    # 5. Date fields validation
    for date_field in ['Due_date', 'Start_Date', 'Created_date']:
        val = api_body.get(date_field)
        if val is not None and val != '':
            val_str = str(val).strip()
            if val_str:
                try:
                    dateutil_parser.parse(val_str)
                except (ValueError, TypeError):
                    reasons.append(f"{date_field} must be a valid date format")
                    
    return reasons

@app.route('/api/validate-import', methods=['POST'])
@app.route('/validate-import', methods=['POST'])
def validate_import():
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({"error": "No data provided"}), 400

        data = payload.get('data', [])
        mapping = payload.get('mapping', {})

        if not isinstance(data, list):
            return jsonify({"error": "Data must be a list"}), 400
        if not isinstance(mapping, dict):
            return jsonify({"error": "Mapping must be a dictionary"}), 400

        rows_result = []
        valid_count = 0
        invalid_count = 0

        mandatory_fields = payload.get('mandatory_fields', [])

        # Build reverse mapping: orcanos_field -> excel_col
        for row_idx, row in enumerate(data, 1):
            project_config = payload.get('projectConfig', {})
            orcanos_fields = payload.get('orcanosFields', [])
            custom_fields = [
                f for f in orcanos_fields
                if isinstance(f.get('name'), str) and
                f['name'].startswith('CS') and
                f['name'][2:].isdigit()
            ]
            custom_field_index = {
                f['ws_add_col_name'].replace('_Name', '_value'): idx + 1
                for idx, f in enumerate(custom_fields)
            }

            # Resolve all mapped fields
            resolved_mapping = {}
            for orcanos_field, parts in mapping.items():
                if not isinstance(parts, list) or len(parts) == 0:
                    continue

                resolved_value = ""
                for part in parts:
                    if not isinstance(part, dict):
                        continue
                    part_type = part.get('type')
                    part_val = part.get('value')
                    if part_type == 'column':
                        cell_val = row.get(part_val)
                        if cell_val is not None:
                            resolved_value += str(cell_val)
                    elif part_type == 'text':
                        if part_val is not None:
                            resolved_value += str(part_val)
                resolved_mapping[orcanos_field] = resolved_value

            # Determine path (Add vs Update) based on Object_ID
            object_id_val = resolved_mapping.get('Object_ID', '').strip()
            is_update = bool(object_id_val)

            # Build api_body
            api_body = {}
            api_body['Project_ID'] = int(project_config.get('project_id', 0))
            api_body['Major_Version'] = int(project_config.get('major_version', 0))
            api_body['Minor_Version'] = int(project_config.get('minor_version', 0))
            api_body['Object_Type'] = project_config.get('object_type_label', project_config.get('item_type', ''))

            for orcanos_field, resolved_value in resolved_mapping.items():
                # If this is Add path, skip empty/whitespace values (existing behavior)
                if not is_update:
                    if not resolved_value.strip():
                        continue

                # Check if this is a custom field
                if orcanos_field in custom_field_index:
                    n = custom_field_index[orcanos_field]
                    field_title = next(
                        (f.get('title', '') for f in custom_fields if f['ws_add_col_name'].replace('_Name', '_value') == orcanos_field),
                        ''
                    )
                    api_body[f'CS{n}_Name'] = field_title
                    api_body[f'CS{n}_value'] = resolved_value
                else:
                    api_body[orcanos_field] = resolved_value

            if not api_body.get('Parent_ID'):
                api_body['Parent_ID'] = str(api_body['Project_ID'])

            item_type_code = project_config.get('item_type', '').upper()
            # QW_Add/Update_Defect requires the raw project name (without version suffix) for project resolution
            if item_type_code == 'DEFECT':
                api_body['Project_Name'] = project_config.get('raw_project_name') or project_config.get('project_name', '')
            if is_update and item_type_code in ('T_CASE', 'DEFECT'):
                api_body.pop('Release_Version', None)

            if is_update:
                reasons = []
            else:
                reasons = validate_row(api_body, mandatory_fields)
                field_name_to_title = {
                    f['ws_add_col_name'].replace('_Name', '_value'): f.get('title', f.get('name', ''))
                    for f in orcanos_fields
                }
                reasons = [
                    next((r.replace(fname, title) for fname, title in field_name_to_title.items() if fname in r), r)
                    for r in reasons
                ]

            is_valid = len(reasons) == 0
            if is_valid:
                valid_count += 1
            else:
                invalid_count += 1

            # For DEFECT the display name field is Synopsis; fall back to Object_Name
            display_name = api_body.get('Synopsis', '') if item_type_code == 'DEFECT' else api_body.get('Object_Name', '')
            rows_result.append({
                "row": row_idx,
                "objectName": serialize_value(display_name) or '',
                "objectType": serialize_value(api_body.get('Object_Type', '')) or '',
                "valid": is_valid,
                "reasons": reasons
            })

        return Response(
            json.dumps({
                "totalRows": len(data),
                "validRows": valid_count,
                "invalidRows": invalid_count,
                "rows": rows_result
            }, default=serialize_value),
            mimetype='application/json',
            status=200
        )

    except Exception as e:
        return jsonify({"error": f"Error validating import: {str(e)}"}), 500

@app.route('/api/import', methods=['POST'])
@app.route('/import', methods=['POST'])
def import_data():
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({"error": "No data provided"}), 400
            
        data = payload.get('data', [])
        mapping = payload.get('mapping', {})
        domain = payload.get('domain', '')
        headers = payload.get('headers', {})
        
        if not domain:
            return jsonify({"error": "Domain is required"}), 400
        
        if not isinstance(data, list):
            return jsonify({"error": "Data must be a list"}), 400
        
        if not isinstance(mapping, dict):
            return jsonify({"error": "Mapping must be a dictionary"}), 400

        results = []
        added_count = 0
        updated_count = 0
        failed_count = 0
        skipped_count = 0
        mandatory_fields = payload.get('mandatory_fields', [])

        # Steps import fields (only present for Test Case with steps sheet)
        steps_data       = payload.get('stepsData') or []
        steps_mapping    = payload.get('stepsMapping') or {}
        tc_link_col      = payload.get('testCaseLinkColumn', '')  # column in main sheet
        step_link_col    = payload.get('stepsLinkColumn', '')     # column in steps sheet
        has_steps_import = bool(steps_data and steps_mapping and tc_link_col and step_link_col)
        
        def generate():
            nonlocal added_count, updated_count, failed_count, skipped_count

            def resolve_parts(parts, row):
                """Resolve a mapping parts list against a data row, returning the string value."""
                resolved = ""
                for part in (parts or []):
                    if not isinstance(part, dict):
                        continue
                    if part.get('type') == 'column':
                        cell_val = row.get(part.get('value'))
                        if cell_val is not None:
                            resolved += str(cell_val)
                    elif part.get('type') == 'text':
                        if part.get('value') is not None:
                            resolved += str(part.get('value'))
                return resolved

            for row_idx, row in enumerate(data, 1):
                try:
                    # Resolve all mapped fields
                    resolved_mapping = {}
                    for orcanos_field, parts in mapping.items():
                        if not isinstance(parts, list) or len(parts) == 0:
                            continue

                        resolved_value = ""
                        for part in parts:
                            if not isinstance(part, dict):
                                continue
                            part_type = part.get('type')
                            part_val = part.get('value')
                            if part_type == 'column':
                                cell_val = row.get(part_val)
                                if cell_val is not None:
                                    resolved_value += str(cell_val)
                            elif part_type == 'text':
                                if part_val is not None:
                                    resolved_value += str(part_val)
                        resolved_mapping[orcanos_field] = resolved_value

                    # Determine path (Add vs Update) based on Object_ID
                    object_id_val = resolved_mapping.get('Object_ID', '').strip()
                    is_update = bool(object_id_val)

                    # Build api_body
                    api_body = {}

                    # Auto-inject from projectConfig
                    project_config = payload.get('projectConfig', {})
                    api_body['Project_ID'] = int(project_config.get('project_id', 0))
                    api_body['Major_Version'] = int(project_config.get('major_version', 0))
                    api_body['Minor_Version'] = int(project_config.get('minor_version', 0))
                    api_body['Object_Type'] = project_config.get('object_type_label', project_config.get('item_type', ''))
                    # Get orcanosFields for custom field title lookup
                    orcanos_fields = payload.get('orcanosFields', [])
                    custom_fields = [
                        f for f in orcanos_fields
                        if isinstance(f.get('name'), str) and
                        f['name'].startswith('CS') and
                        f['name'][2:].isdigit()
                    ]

                    # Build a lookup: field name -> CS slot index (1-based)
                    custom_field_index = {
                        f['ws_add_col_name'].replace('_Name', '_value'): idx + 1
                        for idx, f in enumerate(custom_fields)
                    }

                    for orcanos_field, resolved_value in resolved_mapping.items():
                        # If this is Add path, skip empty/whitespace values (existing behavior)
                        if not is_update:
                            if not resolved_value.strip():
                                continue

                        # Check if this is a custom field
                        if orcanos_field in custom_field_index:
                            n = custom_field_index[orcanos_field]
                            field_title = next(
                                (f.get('title', '') for f in custom_fields if f['ws_add_col_name'].replace('_Name', '_value') == orcanos_field),
                                ''
                            )
                            api_body[f'CS{n}_Name'] = field_title
                            api_body[f'CS{n}_value'] = resolved_value
                        else:
                            api_body[orcanos_field] = resolved_value

                    if is_update:
                        try:
                            api_body['Object_ID'] = int(float(object_id_val))
                        except ValueError:
                            api_body['Object_ID'] = object_id_val

                    # Parent_ID defaulting
                    if not api_body.get('Parent_ID'):
                        api_body['Parent_ID'] = str(api_body['Project_ID'])

                    item_type_code = project_config.get('item_type', '').upper()
                    # QW_Add/Update_Defect requires the raw project name (without version suffix) for project resolution
                    if item_type_code == 'DEFECT':
                        api_body['Project_Name'] = project_config.get('raw_project_name') or project_config.get('project_name', '')
                    if is_update and item_type_code in ('T_CASE', 'DEFECT'):
                        api_body.pop('Release_Version', None)
                    
                    # Validate row if Add
                    validation_errors = [] if is_update else validate_row(api_body, mandatory_fields)
                    
                    # For DEFECT the display name field is Synopsis; fall back to Object_Name
                    display_name = api_body.get('Synopsis', '') if item_type_code == 'DEFECT' else api_body.get('Object_Name', '')
                    result = {
                        'row': row_idx,
                        'objectName': serialize_value(display_name),
                        'objectType': serialize_value(api_body.get('Object_Type', '')),
                        'status': 'pending',
                        'objectId': 0,
                        'error': ''
                    }
                    
                    if validation_errors:
                        result['status'] = 'skipped'
                        result['error'] = f"Validation failed: {', '.join(validation_errors)}"
                        skipped_count += 1
                    else:
                        # Call Orcanos API
                        try:
                            if item_type_code == 'DEFECT':
                                url = f"https://{domain}/api/v2/Json/QW_Update_Defect" if is_update else f"https://{domain}/api/v2/Json/QW_Add_Defect"
                            elif is_update:
                                url = f"https://{domain}/api/v2/Json/QW_Update_Object"
                            else:
                                url = f"https://{domain}/api/v2/Json/QW_Add_Object"

                            response = requests.post(url, json=api_body, headers=headers, timeout=30)
                            
                            if response.status_code == 200:
                                response_data, parse_error = parse_orcanos_json(response)
                                if parse_error:
                                    result['status'] = 'failed'
                                    result['error'] = parse_error
                                    failed_count += 1
                                else:
                                    object_id = response_data.get('Data', 0)
                                    fallback_msg = "Object was not updated." if is_update else "Object was not created."
                                    
                                    if isinstance(object_id, dict):
                                        error_info = object_id.get('ErrorInfo', '') or ''
                                        if "There is no row at position 0." in error_info:
                                            error_info = "object does not exist"
                                        result['status'] = 'failed'
                                        result['error'] = error_info if error_info else fallback_msg
                                        failed_count += 1
                                    elif object_id and isinstance(object_id, (int, float)) and int(object_id) > 0:
                                        result['status'] = 'updated' if is_update else 'added'
                                        result['objectId'] = int(object_id)
                                        if is_update:
                                            updated_count += 1
                                        else:
                                            added_count += 1

                                        # ── Import steps if available ──────────────────
                                        if has_steps_import:
                                            resolved_tc_key = str(row.get(tc_link_col, '')).strip()
                                            matching_steps = [
                                                s for s in steps_data
                                                if str(s.get(step_link_col, '')).strip() == resolved_tc_key
                                            ]
                                            steps_total  = len(matching_steps)
                                            steps_added  = 0
                                            steps_failed = 0

                                            for step_row in matching_steps:
                                                step_body = {
                                                    'ItemId':        int(object_id),
                                                    'ObjectType':    'DEFECT' if item_type_code == 'DEFECT' else 'OBJECT',
                                                    'StepNumber':    resolve_parts(steps_mapping.get('StepNumber'),    step_row).strip(),
                                                    'Description':   resolve_parts(steps_mapping.get('Description'),   step_row).strip(),
                                                    'ExpectedValue': resolve_parts(steps_mapping.get('ExpectedValue'), step_row).strip(),
                                                    'LowerLimit':    resolve_parts(steps_mapping.get('LowerLimit'),    step_row).strip(),
                                                    'UpperLimit':    resolve_parts(steps_mapping.get('UpperLimit'),    step_row).strip(),
                                                }
                                                try:
                                                    step_url = f"https://{domain}/api/v2/Json/AddStep"
                                                    step_resp = requests.post(
                                                        step_url, json=step_body,
                                                        headers=headers, timeout=30
                                                    )
                                                    if step_resp.status_code == 200:
                                                        step_data, step_err = parse_orcanos_json(step_resp)
                                                        if step_err or not step_data or not step_data.get('IsSuccess'):
                                                            steps_failed += 1
                                                        else:
                                                            steps_added += 1
                                                    else:
                                                        steps_failed += 1
                                                except Exception:
                                                    steps_failed += 1

                                            result['stepsTotal']  = steps_total
                                            result['stepsAdded']  = steps_added
                                            result['stepsFailed'] = steps_failed
                                    else:
                                        msg = response_data.get('Message', '') or ''
                                        if "There is no row at position 0." in msg:
                                            msg = "object does not exist"
                                        result['status'] = 'failed'
                                        result['error'] = msg if msg else fallback_msg
                                        failed_count += 1
                            else:
                                result['status'] = 'failed'
                                result['error'] = f"API error: {response.status_code} - {response.text[:200]}"
                                failed_count += 1
                        except requests.exceptions.Timeout:
                            result['status'] = 'failed'
                            result['error'] = "Request timeout - API server not responding"
                            failed_count += 1
                        except requests.exceptions.RequestException as e:
                            result['status'] = 'failed'
                            result['error'] = f"Network error: {str(e)}"
                            failed_count += 1
                        except Exception as e:
                            result['status'] = 'failed'
                            result['error'] = str(e)
                            failed_count += 1
                    
                    results.append(result)
                
                except Exception as e:
                    result = {
                        'row': row_idx,
                        'objectName': '',
                        'objectType': '',
                        'status': 'failed',
                        'objectId': 0,
                        'error': f"Error processing row: {str(e)}"
                    }
                    results.append(result)
                    failed_count += 1
                
                # Stream progress
                yield json.dumps({
                    'type': 'progress',
                    'row': row_idx,
                    'total': len(data)
                }, default=serialize_value) + '\n'
            
            # Stream final results
            yield json.dumps({
                'type': 'done',
                'results': results,
                'summary': {
                    'total': len(data),
                    'success': added_count + updated_count,
                    'added': added_count,
                    'updated': updated_count,
                    'failed': failed_count,
                    'skipped': skipped_count
                }
            }, default=serialize_value) + '\n'
        
        return Response(generate(), mimetype='application/x-ndjson')
    
    except Exception as e:
        return jsonify({"error": f"Error processing import: {str(e)}"}), 500
    
@app.route('/health')
def health():
    return {"status": "ok"}

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)